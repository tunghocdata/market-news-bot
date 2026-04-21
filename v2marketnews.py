from __future__ import annotations
"""
v2marketnews.py – News Fetch Pipeline v2 (standalone, không phụ thuộc pipeline hiện tại)

Chạy thử:
  python3 v2marketnews.py --rss-only --no-enrich   # nhanh nhất ~15s
  python3 v2marketnews.py --no-enrich               # RSS + scrape ~30s
  python3 v2marketnews.py                           # full pipeline ~5-8 phút
"""

import asyncio
import hashlib
import json
import logging
import os
import ssl
import sys
import urllib.request
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta
from email.utils import parsedate_to_datetime

import json as _json
import re

import aiohttp
import feedparser
import requests
import urllib3
from bs4 import BeautifulSoup
from dotenv import load_dotenv

try:
    import trafilatura as _trafilatura
    import logging as _lg
    _lg.getLogger("trafilatura").setLevel(_lg.ERROR)
except ImportError:
    _trafilatura = None

load_dotenv()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    level=logging.WARNING,   # chỉ hiển thị WARNING+ từ thư viện ngoài
)
log = logging.getLogger("v2news")
log.setLevel(logging.INFO)

# ── Flags CLI ─────────────────────────────────────────────────────────────────
RSS_ONLY   = "--rss-only"   in sys.argv
NO_ENRICH  = "--no-enrich"  in sys.argv

# --hours N  (mặc định 24)
_hours_arg = next((sys.argv[i+1] for i, a in enumerate(sys.argv)
                   if a == "--hours" and i+1 < len(sys.argv)), "24")
try:
    FILTER_HOURS = int(_hours_arg)
except ValueError:
    FILTER_HOURS = 24

# ── Config ────────────────────────────────────────────────────────────────────
JINA_API_KEY            = os.getenv("JINA_API_KEY", "")
GUARDIAN_API_KEY        = os.getenv("GUARDIAN_API_KEY", "")
FETCH_RSS_WORKERS       = 12
FETCH_SCRAPE_WORKERS    = 6
ENRICH_WORKERS          = 20

# Domains được biết là paywall — skip trafilatura/Jina, dùng RSS summary luôn
_PAYWALL_DOMAINS = frozenset({
    "bloomberg.com", "wsj.com", "ft.com",
    "nytimes.com", "economist.com", "barrons.com",
})
ENRICH_TIMEOUT_TOTAL    = 600    # tối đa 10 phút cho toàn bộ enrich
SUMMARY_MIN_LEN         = 200    # VN bài có summary đủ dài → skip enrich
_VN_TZ = timezone(timedelta(hours=7))

# Meta tags đáng tin cậy cho pub_time (KHÔNG dùng time[datetime] — quá nhiễu sidebar)
_PUBLISH_META_SELECTORS = [
    'meta[property="article:published_time"]',
    'meta[property="og:published_time"]',
    'meta[name="pubdate"]',
    'meta[name="publishdate"]',
    'meta[name="article:published_time"]',
    'meta[itemprop="datePublished"]',
    'meta[name="DC.date"]',
]

SCRAPE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# SSL context bypass (giống v1 — fix Python 3.14 Mac cert issue)
_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode    = ssl.CERT_NONE

CATEGORIES = {
    "chung_khoan":  "Chứng khoán",
    "tai_chinh":    "Tài chính – Ngân hàng",
    "doanh_nghiep": "Doanh nghiệp",
    "bat_dong_san": "Bất động sản",
    "hang_hoa":     "Hàng hóa",
    "vi_mo":        "Kinh tế vĩ mô",
    "thi_truong":   "Thị trường",
    "the_gioi":     "Thế giới",
    "phan_tich":    "Phân tích",
    "markets":      "Markets (QT)",
}

# ── Nguồn quốc tế — RSS ───────────────────────────────────────────────────────
INTL_RSS_SOURCES = [
    # Reuters (qua Google News proxy — bypass anti-bot)
    {"name": "Reuters – Markets",      "url": "https://news.google.com/rss/search?q=site:reuters.com+markets&hl=en&gl=US&ceid=US:en",             "category": "markets",      "region": "intl"},
    {"name": "Reuters – World",        "url": "https://news.google.com/rss/search?q=site:reuters.com+world&hl=en&gl=US&ceid=US:en",               "category": "the_gioi",     "region": "intl"},
    {"name": "Reuters – Business",     "url": "https://news.google.com/rss/search?q=site:reuters.com+business&hl=en&gl=US&ceid=US:en",            "category": "doanh_nghiep", "region": "intl"},
    {"name": "Reuters – Asia",         "url": "https://news.google.com/rss/search?q=site:reuters.com+asia+pacific&hl=en&gl=US&ceid=US:en",        "category": "the_gioi",     "region": "intl"},
    {"name": "Reuters – Commodities",  "url": "https://news.google.com/rss/search?q=site:reuters.com+commodities&hl=en&gl=US&ceid=US:en",         "category": "hang_hoa",     "region": "intl"},
    # Bloomberg (qua Google News proxy)
    {"name": "Bloomberg – Markets",    "url": "https://news.google.com/rss/search?q=site:bloomberg.com+markets&hl=en&gl=US&ceid=US:en",           "category": "markets",      "region": "intl"},
    {"name": "Bloomberg – Economy",    "url": "https://news.google.com/rss/search?q=site:bloomberg.com+economy&hl=en&gl=US&ceid=US:en",           "category": "vi_mo",        "region": "intl"},
    {"name": "Bloomberg – Asia",       "url": "https://news.google.com/rss/search?q=site:bloomberg.com+asia&hl=en&gl=US&ceid=US:en",              "category": "the_gioi",     "region": "intl"},
    {"name": "Bloomberg – Commodities","url": "https://news.google.com/rss/search?q=site:bloomberg.com+commodities&hl=en&gl=US&ceid=US:en",       "category": "hang_hoa",     "region": "intl"},
    {"name": "Bloomberg – Finance",    "url": "https://news.google.com/rss/search?q=site:bloomberg.com+finance&hl=en&gl=US&ceid=US:en",           "category": "tai_chinh",    "region": "intl"},
    # Financial Times (qua Google News proxy — FT direct RSS cần subscription)
    {"name": "FT – Markets",           "url": "https://news.google.com/rss/search?q=site:ft.com+markets&hl=en&gl=US&ceid=US:en",                  "category": "markets",      "region": "intl"},
    {"name": "FT – Economy",           "url": "https://news.google.com/rss/search?q=site:ft.com+economy&hl=en&gl=US&ceid=US:en",                  "category": "vi_mo",        "region": "intl"},
    {"name": "FT – Asia",              "url": "https://news.google.com/rss/search?q=site:ft.com+asia+pacific&hl=en&gl=US&ceid=US:en",             "category": "the_gioi",     "region": "intl"},
    {"name": "FT – Companies",         "url": "https://news.google.com/rss/search?q=site:ft.com+companies&hl=en&gl=US&ceid=US:en",                "category": "doanh_nghiep", "region": "intl"},
    {"name": "FT – Energy",            "url": "https://news.google.com/rss/search?q=site:ft.com+energy&hl=en&gl=US&ceid=US:en",                   "category": "hang_hoa",     "region": "intl"},
    # NYT
    {"name": "NYT – World",            "url": "https://rss.nytimes.com/services/xml/rss/nyt/World.xml",                                           "category": "the_gioi",     "region": "intl"},
    {"name": "NYT – Business",         "url": "https://rss.nytimes.com/services/xml/rss/nyt/Business.xml",                                        "category": "doanh_nghiep", "region": "intl"},
    {"name": "NYT – Economy",          "url": "https://rss.nytimes.com/services/xml/rss/nyt/Economy.xml",                                         "category": "vi_mo",        "region": "intl"},
    # BBC
    {"name": "BBC – Business",         "url": "https://news.google.com/rss/search?q=site:bbc.com+business+finance&hl=en&gl=US&ceid=US:en",       "category": "doanh_nghiep", "region": "intl"},
    {"name": "BBC – World",            "url": "https://feeds.bbci.co.uk/news/world/rss.xml",                                                      "category": "the_gioi",     "region": "intl"},
    {"name": "BBC – Asia",             "url": "https://feeds.bbci.co.uk/news/world/asia/rss.xml",                                                  "category": "the_gioi",     "region": "intl"},
    # WSJ
    {"name": "WSJ – World",            "url": "https://feeds.a.dj.com/rss/RSSWorldNews.xml",                                                      "category": "the_gioi",     "region": "intl"},
    {"name": "WSJ – Markets",          "url": "https://feeds.a.dj.com/rss/RSSMarketsMain.xml",                                                    "category": "markets",      "region": "intl"},
    {"name": "WSJ – Business",         "url": "https://feeds.a.dj.com/rss/WSJcomUSBusiness.xml",                                                  "category": "doanh_nghiep", "region": "intl"},
    # CNBC
    {"name": "CNBC – Markets",        "url": "https://search.cnbc.com/rs/search/combinedcms/view.xml?partnerId=wrss01&id=15839069", "category": "markets",   "region": "intl"},
    # Project Syndicate (analysis/commentary)
    {"name": "Project Syndicate",     "url": "https://www.project-syndicate.org/rss",                                               "category": "phan_tich", "region": "intl"},
    # Guardian RSS đã bỏ — dùng Guardian API (fetch_guardian_api) thay thế
]

# ── Nguồn trong nước — RSS ────────────────────────────────────────────────────
VN_RSS_SOURCES = [
    # CafeF
    {"name": "CafeF – Chứng khoán",      "url": "https://cafef.vn/thi-truong-chung-khoan.rss",            "category": "chung_khoan",  "region": "vn"},
    {"name": "CafeF – Tài chính NH",     "url": "https://cafef.vn/tai-chinh-ngan-hang.rss",               "category": "tai_chinh",    "region": "vn"},
    {"name": "CafeF – Doanh nghiệp",     "url": "https://cafef.vn/doanh-nghiep.rss",                      "category": "doanh_nghiep", "region": "vn"},
    {"name": "CafeF – Vĩ mô",            "url": "https://cafef.vn/vi-mo-dau-tu.rss",                      "category": "vi_mo",        "region": "vn"},
    {"name": "CafeF – Bất động sản",     "url": "https://cafef.vn/bat-dong-san.rss",                      "category": "bat_dong_san", "region": "vn"},
    {"name": "CafeF – Tài chính QT",     "url": "https://cafef.vn/tai-chinh-quoc-te.rss",                 "category": "the_gioi",     "region": "vn"},
    {"name": "CafeF – Smart Money",      "url": "https://cafef.vn/smart-money.rss",                       "category": "phan_tich",    "region": "vn"},
    # VnEconomy
    {"name": "VnEconomy – Tài chính",    "url": "https://vneconomy.vn/tai-chinh.rss",                     "category": "tai_chinh",    "region": "vn"},
    {"name": "VnEconomy – Chứng khoán",  "url": "https://vneconomy.vn/chung-khoan.rss",                   "category": "chung_khoan",  "region": "vn"},
    {"name": "VnEconomy – BĐS",          "url": "https://vneconomy.vn/bat-dong-san.rss",                  "category": "bat_dong_san", "region": "vn"},
    {"name": "VnEconomy – Doanh nghiệp", "url": "https://vneconomy.vn/doanh-nghiep-1.rss",                "category": "doanh_nghiep", "region": "vn"},
    {"name": "VnEconomy – Thị trường",   "url": "https://vneconomy.vn/thi-truong.rss",                    "category": "thi_truong",   "region": "vn"},
    {"name": "VnEconomy – Thế giới",     "url": "https://vneconomy.vn/the-gioi.rss",                      "category": "the_gioi",     "region": "vn"},
    # Enternews (Diễn đàn doanh nghiệp) — RSS có full body text
    {"name": "Enternews – Doanh nghiệp","url": "https://enternews.vn/feed",                                 "category": "doanh_nghiep", "region": "vn"},
    # Người lao động — kinh tế
    {"name": "NLD – Kinh tế",           "url": "https://nld.com.vn/rss/kinh-te.rss",                        "category": "vi_mo",        "region": "vn"},
    # VnExpress (sub-RSS path sai → dùng Google News proxy; BĐS dùng direct)
    {"name": "VnExpress – Kinh doanh",   "url": "https://news.google.com/rss/search?q=site:vnexpress.net+kinh-doanh&hl=vi&gl=VN&ceid=VN:vi",    "category": "doanh_nghiep", "region": "vn"},
    {"name": "VnExpress – Tài chính",    "url": "https://news.google.com/rss/search?q=site:vnexpress.net+tai-chinh&hl=vi&gl=VN&ceid=VN:vi",     "category": "tai_chinh",    "region": "vn"},
    {"name": "VnExpress – Chứng khoán",  "url": "https://news.google.com/rss/search?q=site:vnexpress.net+chung-khoan&hl=vi&gl=VN&ceid=VN:vi",   "category": "chung_khoan",  "region": "vn"},
    {"name": "VnExpress – BĐS",          "url": "https://vnexpress.net/rss/bat-dong-san.rss",                                                     "category": "bat_dong_san", "region": "vn"},
    {"name": "VnExpress – Thế giới",     "url": "https://vnexpress.net/rss/the-gioi.rss",                                                         "category": "the_gioi",     "region": "vn"},
    # Người Quan Sát (Google News proxy — site dùng Cloudflare, URL-encode tiếng Việt)
    {"name": "NQS – Chứng khoán",        "url": "https://news.google.com/rss/search?q=site:nguoiquansat.vn+ch%E1%BB%A9ng+kho%C3%A1n&hl=vi&gl=VN&ceid=VN:vi",      "category": "chung_khoan",  "region": "vn"},
    {"name": "NQS – Bất động sản",       "url": "https://news.google.com/rss/search?q=site:nguoiquansat.vn+b%E1%BA%A5t+%C4%91%E1%BB%99ng+s%E1%BA%A3n&hl=vi&gl=VN&ceid=VN:vi", "category": "bat_dong_san", "region": "vn"},
    {"name": "NQS – Ngân hàng",          "url": "https://news.google.com/rss/search?q=site:nguoiquansat.vn+ng%C3%A2n+h%C3%A0ng&hl=vi&gl=VN&ceid=VN:vi",            "category": "tai_chinh",    "region": "vn"},
    {"name": "NQS – Doanh nghiệp",       "url": "https://news.google.com/rss/search?q=site:nguoiquansat.vn+doanh+nghi%E1%BB%87p&hl=vi&gl=VN&ceid=VN:vi",           "category": "doanh_nghiep", "region": "vn"},
    {"name": "NQS – Vĩ mô",              "url": "https://news.google.com/rss/search?q=site:nguoiquansat.vn+v%C4%A9+m%C3%B4&hl=vi&gl=VN&ceid=VN:vi",                "category": "vi_mo",        "region": "vn"},
]

# ── Nguồn trong nước — HTML scrape (per-category page) ───────────────────────
VN_SCRAPE_SOURCES = [
    # Vietstock (per-category — ổn định hơn scrape trang chủ)
    {"name": "Vietstock – Chứng khoán",   "url": "https://vietstock.vn/chung-khoan.htm",       "selectors": ["h4 a", "a.title-news", "h3 a"],                "base_url": "https://vietstock.vn", "category": "chung_khoan",  "region": "vn"},
    {"name": "Vietstock – Doanh nghiệp",  "url": "https://vietstock.vn/doanh-nghiep.htm",      "selectors": ["h4 a", "a.title-news", "h3 a"],                "base_url": "https://vietstock.vn", "category": "doanh_nghiep", "region": "vn"},
    {"name": "Vietstock – Bất động sản",  "url": "https://vietstock.vn/bat-dong-san.htm",      "selectors": ["h4 a", "a.title-news", "h3 a"],                "base_url": "https://vietstock.vn", "category": "bat_dong_san", "region": "vn"},
    {"name": "Vietstock – Tài chính",     "url": "https://vietstock.vn/tai-chinh.htm",         "selectors": ["h4 a", "a.title-news", "h3 a"],                "base_url": "https://vietstock.vn", "category": "tai_chinh",    "region": "vn"},
    {"name": "Vietstock – Hàng hóa",      "url": "https://vietstock.vn/hang-hoa.htm",          "selectors": ["h4 a", "a.title-news", "h3 a"],                "base_url": "https://vietstock.vn", "category": "hang_hoa",     "region": "vn"},
    {"name": "Vietstock – Kinh tế",       "url": "https://vietstock.vn/kinh-te.htm",           "selectors": ["h4 a", "a.title-news", "h3 a"],                "base_url": "https://vietstock.vn", "category": "vi_mo",        "region": "vn"},
    {"name": "Vietstock – Thế giới",      "url": "https://vietstock.vn/the-gioi.htm",          "selectors": ["h4 a", "a.title-news", "h3 a"],                "base_url": "https://vietstock.vn", "category": "the_gioi",     "region": "vn"},
    # VietnamBiz (3/6 category dùng JS render — chỉ giữ 3 category có SSR)
    {"name": "VietnamBiz – Chứng khoán",  "url": "https://vietnambiz.vn/chung-khoan.htm",      "selectors": ["h3 a", "p.story__title a", ".story__title a"], "base_url": "https://vietnambiz.vn", "category": "chung_khoan",  "region": "vn"},
    {"name": "VietnamBiz – Hàng hóa",     "url": "https://vietnambiz.vn/hang-hoa.htm",         "selectors": ["h3 a", "p.story__title a", ".story__title a"], "base_url": "https://vietnambiz.vn", "category": "hang_hoa",     "region": "vn"},
    {"name": "VietnamBiz – Doanh nghiệp", "url": "https://vietnambiz.vn/doanh-nghiep.htm",     "selectors": ["h3 a", "p.story__title a", ".story__title a"], "base_url": "https://vietnambiz.vn", "category": "doanh_nghiep", "region": "vn"},
]


# ─────────────────────────────────────────────────────────────────────────────
# BƯỚC 1 — FETCH RSS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_rss_dt(entry) -> str:
    """Parse RSS datetime → ISO string (VN tz), hoặc rỗng nếu không có."""
    for attr in ("published", "updated"):
        raw = entry.get(attr)
        if raw:
            try:
                dt = parsedate_to_datetime(raw)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt.astimezone(_VN_TZ).isoformat()
            except Exception:
                pass
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# PUB TIME — Extract publish datetime từ HTML (meta tags + JSON-LD)
# Port từ scraper.py (proven trên Vietstock, VietnamBiz, CafeF, VnEconomy)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_dt_str(raw: str) -> datetime | None:
    """Parse chuỗi datetime nhiều format → timezone-aware datetime (VN tz) hoặc None."""
    raw = raw.strip()
    if not raw or len(raw) < 8:
        return None
    # 1. RFC 2822
    try:
        dt = parsedate_to_datetime(raw)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_VN_TZ)
    except Exception:
        pass
    # 2. ISO 8601
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=_VN_TZ)
        return dt.astimezone(_VN_TZ)
    except Exception:
        pass
    # 3. Các format không có tz (assume VN tz)
    for fmt, n in [
        ("%Y-%m-%dT%H:%M:%S", 19), ("%Y-%m-%d %H:%M:%S", 19),
        ("%Y-%m-%d %H:%M", 16),    ("%d/%m/%Y %H:%M:%S", 19),
        ("%d/%m/%Y %H:%M", 16),
    ]:
        try:
            return datetime.strptime(raw[:n], fmt).replace(tzinfo=_VN_TZ)
        except ValueError:
            continue
    return None


def _is_plausible_pub_time(dt: datetime, max_days_ago: int = 7) -> bool:
    """Sanity-check: loại bỏ datetime quá cũ (>7 ngày) hoặc tương lai."""
    now = datetime.now(_VN_TZ)
    return (now - timedelta(days=max_days_ago)) <= dt <= (now + timedelta(hours=2))


def _extract_publish_time(html_text: str) -> str | None:
    """
    Trích xuất thời gian xuất bản thật của bài báo.
    Ưu tiên: Meta tags → JSON-LD. KHÔNG dùng time[datetime] (sidebar noise).
    """
    try:
        soup = BeautifulSoup(html_text, "html.parser")
        # Nguồn 1: Meta tags
        for sel in _PUBLISH_META_SELECTORS:
            el = soup.select_one(sel)
            if not el:
                continue
            raw = (el.get("content") or el.get("datetime") or "").strip()
            if not raw:
                continue
            dt = _parse_dt_str(raw)
            if dt and _is_plausible_pub_time(dt):
                return dt.isoformat()
        # Nguồn 2: JSON-LD (Schema.org)
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = _json.loads(script.string or "")
                nodes = data if isinstance(data, list) else data.get("@graph", [data])
                for node in (nodes if isinstance(nodes, list) else [nodes]):
                    if not isinstance(node, dict):
                        continue
                    raw = node.get("datePublished") or node.get("dateCreated") or ""
                    if raw:
                        dt = _parse_dt_str(str(raw))
                        if dt and _is_plausible_pub_time(dt):
                            return dt.isoformat()
            except Exception:
                continue
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────────────────────
# GUARDIAN API — trả về full body text, không cần enrich
# ─────────────────────────────────────────────────────────────────────────────

# section → category mapping
_GUARDIAN_SECTIONS = [
    ("world",       "the_gioi"),
    ("business",    "doanh_nghiep"),
    ("money",       "tai_chinh"),
    ("environment", "hang_hoa"),
    ("us-news",     "the_gioi"),
    ("global-development", "the_gioi"),
]

_GUARDIAN_API_URL = "https://content.guardianapis.com/search"


def _fetch_guardian_section(section: str, category: str) -> list[dict]:
    """Gọi Guardian API cho 1 section, trả về articles với full body text."""
    if not GUARDIAN_API_KEY:
        return []
    try:
        r = requests.get(
            _GUARDIAN_API_URL,
            params={
                "api-key":     GUARDIAN_API_KEY,
                "section":     section,
                "show-fields": "headline,trailText,bodyText",
                "page-size":   50,
                "order-by":    "newest",
            },
            timeout=12,
        )
        r.raise_for_status()
        results = r.json().get("response", {}).get("results", [])
        articles = []
        for item in results:
            url   = item.get("webUrl", "")
            title = item.get("webTitle", "").strip()
            if not url or not title:
                continue
            fields    = item.get("fields", {})
            body_text = (fields.get("bodyText") or "").strip()
            trail     = (fields.get("trailText") or "").strip()
            pub_date  = item.get("webPublicationDate", "")
            # Convert ISO datetime (UTC) → VN tz
            try:
                from datetime import datetime as _dt
                dt = _dt.fromisoformat(pub_date.replace("Z", "+00:00"))
                pub_date = dt.astimezone(_VN_TZ).isoformat()
            except Exception:
                pass
            articles.append({
                "id":            hashlib.md5(url.encode()).hexdigest(),
                "title":         title,
                "url":           url,
                "summary":       trail[:600],
                "published":     pub_date,
                "datetime":      pub_date,
                "source":        f"Guardian – {section.replace('-',' ').title()}",
                "category":      category,
                "region":        "intl",
                "lang":          "en",
                "full_text":     body_text[:3000],
                "enrich_method": "guardian_api" if body_text else "rss_summary_fallback",
            })
        log.info("  Guardian API %-20s : %2d bài", section, len(articles))
        return articles
    except Exception as e:
        log.warning("Guardian API [%s] lỗi: %s", section, e)
        return []


def fetch_guardian_api() -> list[dict]:
    """Fetch tất cả Guardian sections song song."""
    if not GUARDIAN_API_KEY:
        return []
    all_articles: list[dict] = []
    with ThreadPoolExecutor(max_workers=6) as ex:
        futures = {ex.submit(_fetch_guardian_section, sec, cat): sec
                   for sec, cat in _GUARDIAN_SECTIONS}
        for fut in as_completed(futures):
            try:
                all_articles.extend(fut.result())
            except Exception as e:
                log.warning("Guardian API lỗi: %s", e)
    return all_articles


def _feedparser_fetch(url: str):
    """feedparser với SSL bypass. Google News proxy = 15s, direct RSS = 8s."""
    timeout = 15 if "news.google.com" in url else 8
    try:
        req  = urllib.request.Request(url, headers={"User-Agent": "feedparser/6.0"})
        resp = urllib.request.urlopen(req, context=_SSL_CTX, timeout=timeout)
        return feedparser.parse(resp.read())
    except Exception:
        return feedparser.parse("")


def _rss_extract_text(entry) -> tuple[str, str]:
    """Trả về (summary, full_body) từ RSS entry.
    full_body != "" khi feed cung cấp content:encoded (e.g. Enternews, Project Syndicate).
    """
    # content:encoded → full body
    full_body = ""
    if entry.get("content"):
        raw = entry["content"][0].get("value", "")
        if raw:
            try:
                full_body = BeautifulSoup(raw, "html.parser").get_text(" ", strip=True)
            except Exception:
                full_body = raw
    # summary / description
    raw_sum = entry.get("summary") or entry.get("description") or ""
    try:
        summary = BeautifulSoup(raw_sum, "html.parser").get_text(" ", strip=True)
    except Exception:
        summary = raw_sum
    return summary[:600], full_body[:3000]


def parse_rss_source(source: dict) -> list[dict]:
    feed     = _feedparser_fetch(source["url"])
    articles = []
    for entry in feed.entries:
        url   = entry.get("link", "").strip()
        title = entry.get("title", "").strip()
        if not url or not title or len(title) < 8:
            continue
        summary, full_body = _rss_extract_text(entry)
        enrich_method = "rss_full_body" if len(full_body) >= 200 else ""
        pub_dt = _parse_rss_dt(entry)
        articles.append({
            "id":            hashlib.md5(url.encode()).hexdigest(),
            "title":         title,
            "url":           url,
            "summary":       summary,
            "published":     pub_dt,
            "datetime":      pub_dt,
            "source":        source["name"],
            "category":      source["category"],
            "region":        source["region"],
            "lang":          "vi" if source["region"] == "vn" else "en",
            "full_text":     full_body,
            "enrich_method": enrich_method,
        })
    log.info("  RSS %-35s : %2d bài", source["name"], len(articles))
    return articles


def fetch_all_rss(sources: list[dict]) -> list[dict]:
    all_articles: list[dict] = []
    with ThreadPoolExecutor(max_workers=FETCH_RSS_WORKERS) as ex:
        futures = {ex.submit(parse_rss_source, s): s["name"] for s in sources}
        for fut in as_completed(futures):
            try:
                all_articles.extend(fut.result())
            except Exception as e:
                log.warning("RSS lỗi [%s]: %s", futures[fut], e)
    return all_articles


# ─────────────────────────────────────────────────────────────────────────────
# BƯỚC 2 — SCRAPE HTML (VN per-category)
# ─────────────────────────────────────────────────────────────────────────────

def scrape_vn_source(source: dict) -> list[dict]:
    try:
        # Per-call session tránh thread-safety issue với shared session
        sess = requests.Session()
        sess.headers.update(SCRAPE_HEADERS)
        r = sess.get(source["url"], timeout=(3, 8), verify=False)
        r.raise_for_status()
        soup    = BeautifulSoup(r.text, "html.parser")
        base    = source["base_url"].rstrip("/")
        seen    = set()
        articles: list[dict] = []

        for selector in source["selectors"]:
            for tag in soup.select(selector):
                href  = tag.get("href", "").strip()
                title = tag.get_text(strip=True)
                if not href or not title or len(title) < 10:
                    continue
                if not href.startswith("http"):
                    href = base + ("" if href.startswith("/") else "/") + href
                from urllib.parse import urlparse as _up
                path = _up(href).path.rstrip("/")
                segs = [s for s in path.split("/") if s]
                is_article = (
                    path.endswith((".htm", ".html", ".chn"))
                    or (len(segs) >= 2 and len(segs[-1]) > 10)
                )
                if not is_article or href in seen:
                    continue
                seen.add(href)

                # Tìm published date từ <time datetime="..."> gần nhất
                pub_date = ""
                parent = tag.parent
                for _ in range(4):
                    if parent is None:
                        break
                    time_tag = parent.find("time")
                    if time_tag and time_tag.get("datetime"):
                        try:
                            dt = datetime.fromisoformat(
                                time_tag["datetime"].replace("Z", "+00:00")
                            )
                            if dt.tzinfo is None:
                                dt = dt.replace(tzinfo=timezone.utc)
                            pub_date = dt.astimezone(_VN_TZ).isoformat()
                            break
                        except Exception:
                            pass
                    parent = parent.parent

                articles.append({
                    "id":            hashlib.md5(href.encode()).hexdigest(),
                    "title":         title,
                    "url":           href,
                    "summary":       "",
                    "published":     pub_date,
                    "datetime":      pub_date,
                    "source":        source["name"],
                    "category":      source["category"],
                    "region":        source["region"],
                    "lang":          "vi",
                    "full_text":     "",
                    "enrich_method": "scraped",
                })
            if articles:
                break  # dùng selector đầu tiên có kết quả

        articles = articles[:50]
        log.info("  Scrape %-32s : %2d bài", source["name"], len(articles))
        return articles
    except Exception as e:
        log.warning("Scrape lỗi [%s]: %s", source["name"], e)
        return []


_SCRAPE_GLOBAL_TIMEOUT = 90  # hard ceiling: toàn bộ HTML scrape xong trong 90s

def fetch_all_scrape(sources: list[dict]) -> list[dict]:
    all_articles: list[dict] = []
    with ThreadPoolExecutor(max_workers=FETCH_SCRAPE_WORKERS) as ex:
        futures = {ex.submit(scrape_vn_source, s): s["name"] for s in sources}
        try:
            for fut in as_completed(futures, timeout=_SCRAPE_GLOBAL_TIMEOUT):
                try:
                    all_articles.extend(fut.result())
                except Exception as e:
                    log.warning("Scrape lỗi: %s", e)
        except TimeoutError:
            log.warning("⚠️ Scrape timeout toàn cục (%ds) — giữ %d bài hiện có",
                        _SCRAPE_GLOBAL_TIMEOUT, len(all_articles))
    return all_articles


# ─────────────────────────────────────────────────────────────────────────────
# BƯỚC 3 — DEDUPLICATE
# ─────────────────────────────────────────────────────────────────────────────

def deduplicate(articles: list[dict]) -> list[dict]:
    seen_urls   = set()
    seen_titles = set()
    unique: list[dict] = []
    for a in articles:
        url_key   = a["url"].split("?")[0].rstrip("/").lower()
        # Normalize title: bỏ dấu câu, collapse spaces → giảm false negative dedup
        title_key = re.sub(r"[^\w\s]", "", a["title"].lower())
        title_key = re.sub(r"\s+", " ", title_key).strip()[:50]
        if url_key in seen_urls or title_key in seen_titles:
            continue
        seen_urls.add(url_key)
        seen_titles.add(title_key)
        unique.append(a)
    return unique


# ─────────────────────────────────────────────────────────────────────────────
# BƯỚC 4 — ENRICH full_text
# Tier 1: trafilatura (free, local)
# Tier 2: Jina AI (optional, paywall bypass)
# Tier 3: RSS summary fallback
# ─────────────────────────────────────────────────────────────────────────────

async def _fetch_pub_time(session: aiohttp.ClientSession, url: str) -> str | None:
    """Lightweight fetch chỉ để lấy pub_time (dùng cho bài scrape thiếu ngày)."""
    try:
        async with session.get(
            url,
            headers=SCRAPE_HEADERS,
            timeout=aiohttp.ClientTimeout(total=6),
            ssl=False,
        ) as r:
            if r.status == 200:
                return _extract_publish_time(await r.text(errors="replace"))
    except Exception:
        pass
    return None


async def _enrich_one(session: aiohttp.ClientSession, article: dict) -> dict:
    # Đã có full_text đủ dài (Guardian API) → skip enrich
    if len(article.get("full_text", "")) >= 200:
        return article

    # VN bài RSS có summary đủ dài → skip enrich
    if article["region"] == "vn" and len(article.get("summary", "")) >= SUMMARY_MIN_LEN:
        article["full_text"]     = article["summary"]
        article["enrich_method"] = "rss_summary"
        return article

    url = article.get("url", "")
    if not url:
        article["full_text"]     = article.get("summary", "")
        article["enrich_method"] = "rss_summary_fallback"
        return article

    # Paywall domains — skip enrich, dùng RSS summary ngay
    from urllib.parse import urlparse as _up
    _domain = _up(url).netloc.lstrip("www.")
    if any(d in _domain for d in _PAYWALL_DOMAINS):
        article["full_text"]     = article.get("summary", "")
        article["enrich_method"] = "rss_only"
        return article

    # Tier 1: trafilatura — fetch_url trong thread pool; đồng thời extract pub_time
    if _trafilatura:
        try:
            def _traf() -> tuple[str, str | None]:
                downloaded = _trafilatura.fetch_url(url)
                text = _trafilatura.extract(
                    downloaded,
                    include_comments=False,
                    include_tables=False,
                    no_fallback=False,
                    output_format="txt",
                ) or ""
                pub_time = _extract_publish_time(downloaded) if downloaded else None
                return text, pub_time
            text, pub_time = await asyncio.wait_for(
                asyncio.to_thread(_traf),
                timeout=10,
            )
            if text and len(text.strip()) > 200:
                article["full_text"]     = text[:3000]
                article["enrich_method"] = "trafilatura"
                if not article.get("published") and pub_time:
                    article["published"] = pub_time
                return article
        except Exception:
            pass

    # Tier 2: Jina AI (nếu có API key)
    if JINA_API_KEY:
        try:
            async with session.get(
                f"https://r.jina.ai/{url}",
                headers={
                    "Authorization": f"Bearer {JINA_API_KEY}",
                    "X-Return-Format": "text",
                    "X-Remove-Selector": "header,footer,nav,.ads,.sidebar,.related,script,.comment",
                    "X-With-Generated-Alt": "false",
                    "X-Timeout": "10",
                },
                timeout=aiohttp.ClientTimeout(total=12),
                ssl=False,
            ) as r:
                if r.status == 200:
                    text = (await r.text())[:3000]
                    # Lọc Jina error/stub responses
                    _bad = (
                        "returned error" in text
                        or "Markdown Content:\nundefined" in text
                        or "X-Error" in text
                        or "paywall" in text.lower()
                        or "subscribe to read" in text.lower()
                        or "access denied" in text.lower()
                    )
                    if not _bad and len(text.strip()) > 300:
                        article["full_text"]     = text
                        article["enrich_method"] = "jina"
                        # Bài scrape VN không có ngày → fetch nhẹ để lấy pub_time
                        if (not article.get("published")
                                and article.get("enrich_method") == "scraped"):
                            pub_time = await _fetch_pub_time(session, url)
                            if pub_time:
                                article["published"] = pub_time
                        return article
        except Exception:
            pass

    # Tier 3: RSS summary fallback — chỉ fetch pub_time cho bài scrape VN thiếu ngày
    if (not article.get("published")
            and article.get("enrich_method") == "scraped"):   # chỉ Vietstock / VietnamBiz
        pub_time = await _fetch_pub_time(session, url)
        if pub_time:
            article["published"] = pub_time
    article["full_text"]     = article.get("summary", "")
    article["enrich_method"] = "rss_summary_fallback"
    return article


async def enrich_all(articles: list[dict]) -> list[dict]:
    sem = asyncio.Semaphore(ENRICH_WORKERS)

    async def bounded(session: aiohttp.ClientSession, a: dict) -> dict:
        async with sem:
            return await _enrich_one(session, a)

    async with aiohttp.ClientSession() as session:
        tasks = [bounded(session, a) for a in articles]
        try:
            results = await asyncio.wait_for(
                asyncio.gather(*tasks, return_exceptions=True),
                timeout=ENRICH_TIMEOUT_TOTAL,
            )
        except asyncio.TimeoutError:
            print("⚠️  Enrich timeout toàn cục — lấy kết quả hiện có")
            results = articles

    return [r for r in results if isinstance(r, dict)]


# ─────────────────────────────────────────────────────────────────────────────
# FILTER & EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def filter_recent(articles: list[dict], hours: int = 24) -> tuple[list[dict], int]:
    """
    Chỉ giữ bài có published trong vòng `hours` giờ gần nhất.
    Bài không có published date → giữ lại (không thể xác định tuổi).
    Trả về (filtered_list, n_dropped).
    """
    cutoff = datetime.now(_VN_TZ) - timedelta(hours=hours)
    kept, dropped = [], 0
    for a in articles:
        pub = a.get("published", "")
        if not pub:
            dropped += 1     # không có ngày → bỏ (không xác định được tuổi bài)
            continue
        try:
            dt = datetime.fromisoformat(pub)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=_VN_TZ)
            if dt >= cutoff:
                kept.append(a)
            else:
                dropped += 1
        except Exception:
            kept.append(a)   # parse lỗi → giữ
    return kept, dropped


def _fmt_pub(pub: str) -> str:
    """ISO datetime → 'HH:MM DD/MM' (VN tz)."""
    if not pub:
        return ""
    try:
        dt = datetime.fromisoformat(pub)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=_VN_TZ)
        return dt.astimezone(_VN_TZ).strftime("%H:%M %d/%m")
    except Exception:
        return pub[:16]


# Màu nền cho từng category (hex, không có #)
_CAT_COLORS: dict[str, str] = {
    "chung_khoan":  "D6E4FF",   # xanh dương nhạt — chứng khoán
    "tai_chinh":    "E6D5F7",   # tím nhạt       — tài chính/ngân hàng
    "doanh_nghiep": "FCE4EC",   # hồng nhạt      — doanh nghiệp
    "bat_dong_san": "FFF8E1",   # vàng nhạt      — bất động sản
    "hang_hoa":     "FBE9E7",   # cam nhạt       — hàng hóa
    "vi_mo":        "E8F5E9",   # xanh lá nhạt   — kinh tế vĩ mô
    "thi_truong":   "F3E5F5",   # tím lạt        — thị trường
    "the_gioi":     "E0F7FA",   # cyan nhạt      — thế giới
    "phan_tich":    "FFFDE7",   # vàng chanh     — phân tích
    "markets":      "E3F2FD",   # trời nhạt      — markets QT
}
_DEFAULT_COLOR = "F5F5F5"

# Emoji prefix cho category label
_CAT_EMOJI: dict[str, str] = {
    "chung_khoan":  "📈",
    "tai_chinh":    "🏦",
    "doanh_nghiep": "🏢",
    "bat_dong_san": "🏠",
    "hang_hoa":     "⚡",
    "vi_mo":        "📊",
    "thi_truong":   "🔄",
    "the_gioi":     "🌍",
    "phan_tich":    "🔍",
    "markets":      "📉",
}


# Tên file master duy nhất (tích lũy qua các ngày)
MASTER_EXCEL_FILE = "market_news_db.xlsx"


def append_to_master_excel(new_articles: list[dict], path: str) -> int:
    """
    Mở file master Excel (tạo mới nếu chưa có).
    - Sheet "📋 Dữ liệu": append bài mới (dedup theo URL), sort mới nhất trên cùng.
    - Sheet "📊 Dashboard": tái tạo hoàn toàn từ toàn bộ dữ liệu.
    Trả về số bài mới được thêm vào.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DATA_SHEET = "📋 Dữ liệu"
    DASH_SHEET = "📊 Dashboard"

    # ── Styles ──────────────────────────────────────────────────────────────
    HEADER_BG   = PatternFill("solid", fgColor="1565C0")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SECT_FONT   = Font(bold=True, size=11, color="1565C0")
    LINK_FONT   = Font(color="1565C0", underline="single", size=10)
    WRAP        = Alignment(wrap_text=True, vertical="top")
    TOP         = Alignment(vertical="top")
    CENTER      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="DDDDDD")
    BORDER      = Border(bottom=thin)
    NUM_FILL    = PatternFill("solid", fgColor="EFF3FF")

    # Columns: (tên, width, wrap?)
    COLS = [
        ("Ngày",     10, False),
        ("Giờ",       7, False),
        ("Khu vực",   8, False),
        ("Nguồn",    22, False),
        ("Nhóm",     22, False),
        ("Tiêu đề",  55, True),
        ("Nội dung", 48, True),
        ("Link",      7, False),
    ]
    URL_COL = 9   # cột ẩn lưu URL gốc để dedup lần sau

    # ── 1. Load existing rows từ file cũ ────────────────────────────────────
    existing_rows: list[dict] = []
    existing_urls: set[str]   = set()

    if os.path.exists(path):
        try:
            wb_old = load_workbook(path, read_only=True, data_only=True)
            if DATA_SHEET in wb_old.sheetnames:
                for row in wb_old[DATA_SHEET].iter_rows(min_row=2, values_only=True):
                    url = row[URL_COL - 1] if len(row) >= URL_COL else ""
                    if not url:
                        continue
                    url_key = str(url).split("?")[0].rstrip("/").lower()
                    existing_urls.add(url_key)
                    region_lbl = str(row[2] or "")
                    cat_lbl    = str(row[4] or "")
                    # Reconstruct category key from emoji prefix
                    cat_key = next(
                        (k for k, v in CATEGORIES.items()
                         if v in cat_lbl or _CAT_EMOJI.get(k, "?") in cat_lbl),
                        "",
                    )
                    existing_rows.append({
                        "_date":     str(row[0] or ""),
                        "_time":     str(row[1] or ""),
                        "_region":   region_lbl,
                        "source":    str(row[3] or ""),
                        "_cat_lbl":  cat_lbl,
                        "title":     str(row[5] or ""),
                        "_content":  str(row[6] or ""),
                        "_url":      str(url),
                        "_sort_key": str(row[0] or "") + str(row[1] or ""),
                        "category":  cat_key,
                        "region":    "vn" if "🇻🇳" in region_lbl else "intl",
                    })
            wb_old.close()
        except Exception as e:
            log.warning("Không đọc được file cũ (%s): %s", os.path.basename(path), e)

    # ── 2. Lọc bài mới (dedup theo URL) ─────────────────────────────────────
    added: list[dict] = []
    for a in new_articles:
        url_key = a["url"].split("?")[0].rstrip("/").lower()
        if url_key not in existing_urls:
            existing_urls.add(url_key)
            added.append(a)

    # ── 3. Convert bài mới → row dict ────────────────────────────────────────
    def _art_to_row(a: dict) -> dict:
        pub = a.get("published", "")
        try:
            dt = datetime.fromisoformat(pub)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=_VN_TZ)
            dt = dt.astimezone(_VN_TZ)
            date_str = dt.strftime("%d/%m/%Y")
            time_str = dt.strftime("%H:%M")
            sort_key = dt.isoformat()
        except Exception:
            date_str = ""
            time_str = ""
            sort_key = pub or "0000"
        cat        = a.get("category", "")
        region     = a.get("region", "")
        return {
            "_date":    date_str,
            "_time":    time_str,
            "_region":  "🇻🇳 VN" if region == "vn" else "🌍 QT",
            "source":   a.get("source", ""),
            "_cat_lbl": f"{_CAT_EMOJI.get(cat,'')}{CATEGORIES.get(cat, cat)}",
            "title":    a.get("title", ""),
            "_content": (a.get("full_text") or a.get("summary") or "")[:300],
            "_url":     a.get("url", ""),
            "_sort_key": sort_key,
            "category": cat,
            "region":   region,
        }

    new_rows = [_art_to_row(a) for a in added]

    # ── 4. Gộp + sort mới nhất trên cùng ────────────────────────────────────
    all_rows = sorted(
        new_rows + existing_rows,
        key=lambda r: r.get("_sort_key", r["_date"] + r["_time"]),
        reverse=True,
    )

    if not all_rows:
        print("⚠️  Không có dữ liệu để ghi.")
        return 0

    # ── 5. Build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet Dữ liệu ────────────────────────────────────────────────────────
    ws = wb.create_sheet(DATA_SHEET)

    for col_idx, (col_name, col_w, _) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col_idx)
        c.value = col_name
        c.fill  = HEADER_BG
        c.font  = HEADER_FONT
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    # Cột ẩn URL (col 9) — dùng để dedup khi đọc lại
    ws.cell(row=1, column=URL_COL).value = "_url"
    ws.column_dimensions[get_column_letter(URL_COL)].width = 0.5
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    for row_idx, r in enumerate(all_rows, 2):
        cat      = r.get("category", "")
        row_fill = PatternFill("solid", fgColor=_CAT_COLORS.get(cat, _DEFAULT_COLOR))
        vals = [
            r["_date"], r["_time"], r["_region"],
            r["source"], r["_cat_lbl"], r["title"], r["_content"],
            None,          # Link col
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value     = val
            c.fill      = row_fill
            c.border    = BORDER
            c.alignment = WRAP if COLS[col_idx - 1][2] else TOP

        url = r.get("_url", "")
        lc  = ws.cell(row=row_idx, column=8)
        lc.fill = row_fill
        lc.border = BORDER
        if url:
            lc.value     = "Đọc →"
            lc.hyperlink = url
            lc.font      = LINK_FONT
            lc.alignment = CENTER

        ws.cell(row=row_idx, column=URL_COL).value = url
        ws.row_dimensions[row_idx].height = 45 if r.get("_content", "").strip() else 20

    # ── Sheet Dashboard ───────────────────────────────────────────────────────
    ws_d = wb.create_sheet(DASH_SHEET)

    from collections import Counter
    date_vn   = Counter()
    date_intl = Counter()
    for r in all_rows:
        d = r.get("_date") or "—"
        if r.get("region") == "vn":
            date_vn[d] += 1
        else:
            date_intl[d] += 1
    all_dates  = sorted(set(date_vn) | set(date_intl), reverse=True)
    cat_cnt    = Counter(r.get("category", "") for r in all_rows)
    src_cnt    = Counter(r.get("source", "")   for r in all_rows)
    total      = len(all_rows)
    now_str    = datetime.now(_VN_TZ).strftime("%H:%M %d/%m/%Y")

    ws_d.merge_cells("A1:L1")
    t = ws_d["A1"]
    t.value = (
        f"📊 DASHBOARD — Cập nhật: {now_str}   |   "
        f"Tổng: {total} bài  •  {len(all_dates)} ngày  •  "
        f"🇻🇳 VN: {sum(date_vn.values())}  •  🌍 QT: {sum(date_intl.values())}"
    )
    t.font      = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws_d.row_dimensions[1].height = 28

    def _tbl(start_row: int, start_col: int, title: str,
             headers: list[str], rows: list[tuple]) -> int:
        """Vẽ bảng thống kê, trả về row tiếp theo sau bảng."""
        nc = len(headers)
        ws_d.merge_cells(
            start_row=start_row, start_column=start_col,
            end_row=start_row,   end_column=start_col + nc - 1,
        )
        c = ws_d.cell(row=start_row, column=start_col)
        c.value = title
        c.font  = SECT_FONT
        r = start_row + 1
        for i, h in enumerate(headers):
            c = ws_d.cell(row=r, column=start_col + i)
            c.value = h
            c.fill  = HEADER_BG
            c.font  = HEADER_FONT
            c.alignment = CENTER
        r += 1
        for row_data in rows:
            for i, val in enumerate(row_data):
                c = ws_d.cell(row=r, column=start_col + i)
                c.value     = val
                c.fill      = NUM_FILL
                c.alignment = CENTER if i > 0 else TOP
            r += 1
        # Tổng cộng
        ws_d.cell(row=r, column=start_col).value = "Tổng cộng"
        ws_d.cell(row=r, column=start_col).font  = Font(bold=True)
        ws_d.cell(row=r, column=start_col + nc - 2).value = total
        ws_d.cell(row=r, column=start_col + nc - 2).font  = Font(bold=True)
        ws_d.cell(row=r, column=start_col + nc - 2).fill  = NUM_FILL
        ws_d.cell(row=r, column=start_col + nc - 2).alignment = CENTER
        return r + 3

    # Bảng 1: Theo Ngày (cột A–D)
    date_rows = [
        (d, date_vn.get(d, 0), date_intl.get(d, 0),
         date_vn.get(d, 0) + date_intl.get(d, 0))
        for d in all_dates
    ]
    _tbl(3, 1, "📅 Theo Ngày", ["Ngày", "🇻🇳 VN", "🌍 QT", "Tổng"], date_rows)

    # Bảng 2: Theo Nhóm (cột F–H)
    cat_rows = [
        (f"{_CAT_EMOJI.get(k,'')}{CATEGORIES.get(k, k)}", v,
         f"{v * 100 / total:.1f}%" if total else "-")
        for k, v in cat_cnt.most_common()
    ]
    _tbl(3, 6, "📂 Theo Nhóm", ["Nhóm", "Bài", "%"], cat_rows)

    # Bảng 3: Theo Nguồn top 25 (cột J–L)
    src_rows = [
        (name, n, f"{n * 100 / total:.1f}%" if total else "-")
        for name, n in src_cnt.most_common(25)
    ]
    _tbl(3, 10, "📰 Theo Nguồn (top 25)", ["Nguồn", "Bài", "%"], src_rows)

    for col, w in [
        (1, 13), (2, 7), (3, 7), (4, 8), (5, 2),
        (6, 28), (7, 7), (8, 7), (9, 2),
        (10, 28), (11, 7), (12, 7),
    ]:
        ws_d.column_dimensions[get_column_letter(col)].width = w

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.active = wb[DASH_SHEET]
    try:
        wb.save(path)
    except PermissionError:
        print(
            f"❌  Lỗi: '{os.path.basename(path)}' đang mở trong Excel. "
            "Đóng file trước rồi chạy lại."
        )
        return 0

    return len(added)


def export_excel(articles: list[dict], path: str) -> None:
    """Xuất Excel đẹp với 4 sheets: VN | QT | Tất cả | Tổng quan."""
    # Kept for backward-compat — internally delegates to append_to_master_excel
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    from collections import Counter

    # ── Chuẩn bị dữ liệu ─────────────────────────────────────────────────────
    def _sort_key(a: dict) -> str:
        return a.get("published", "") or "0000"

    arts = sorted(articles, key=_sort_key, reverse=True)   # mới nhất trước
    vn_arts   = [a for a in arts if a.get("region") == "vn"]
    intl_arts = [a for a in arts if a.get("region") == "intl"]

    # ── Styles dùng chung ─────────────────────────────────────────────────────
    HEADER_BG   = PatternFill("solid", fgColor="1565C0")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT  = Font(bold=True, size=13)
    LINK_FONT   = Font(color="1565C0", underline="single", size=10)
    WRAP        = Alignment(wrap_text=True, vertical="top")
    CENTER      = Alignment(horizontal="center", vertical="center")
    NOWRAP_TOP  = Alignment(vertical="top")
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(bottom=thin)

    # Cột: (tên, width, wrap?)
    COLS = [
        ("Thời gian",  13, False),
        ("Nguồn",      18, False),
        ("Nhóm",       22, False),
        ("Tiêu đề",    55, True),
        ("Nội dung",   48, True),
        ("Link",        7, False),
    ]

    def _write_sheet(ws, data: list[dict], title: str) -> None:
        """Viết một sheet data bài báo."""
        # Row 1: tiêu đề sheet
        ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
        cell = ws["A1"]
        cell.value = title
        cell.font  = TITLE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # Row 2: headers
        for col_idx, (col_name, col_w, _) in enumerate(COLS, 1):
            c = ws.cell(row=2, column=col_idx)
            c.value     = col_name
            c.fill      = HEADER_BG
            c.font      = HEADER_FONT
            c.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_w
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = "A3"   # đóng băng header

        # Rows dữ liệu
        for row_idx, a in enumerate(data, 3):
            cat      = a.get("category", "")
            cat_lbl  = f"{_CAT_EMOJI.get(cat,'')}{CATEGORIES.get(cat, cat)}"
            content  = (a.get("full_text") or a.get("summary") or "")[:300]
            fill_hex = _CAT_COLORS.get(cat, _DEFAULT_COLOR)
            row_fill = PatternFill("solid", fgColor=fill_hex)

            row_vals = [
                _fmt_pub(a.get("published", "")),
                a.get("source", ""),
                cat_lbl,
                a.get("title", ""),
                content,
                None,  # link viết riêng bên dưới
            ]
            for col_idx, val in enumerate(row_vals, 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.value     = val
                c.fill      = row_fill
                c.border    = BORDER
                _, _, wrap  = COLS[col_idx - 1]
                c.alignment = WRAP if wrap else NOWRAP_TOP

            # Cột Link (col 6) — hyperlink "Đọc →"
            url = a.get("url", "")
            lc  = ws.cell(row=row_idx, column=6)
            lc.fill = row_fill
            lc.border = BORDER
            if url:
                lc.value     = "Đọc →"
                lc.hyperlink = url
                lc.font      = LINK_FONT
                lc.alignment = CENTER
            # Chiều cao hàng tự động (openpyxl không tự fit, đặt mặc định đủ dùng)
            ws.row_dimensions[row_idx].height = 45 if content else 18

    # ── Tổng quan sheet ───────────────────────────────────────────────────────
    def _write_summary(ws, all_data: list[dict]) -> None:
        SECT_FONT = Font(bold=True, size=12, color="1565C0")
        NUM_FILL  = PatternFill("solid", fgColor="EFF3FF")

        def _section(start_row: int, title: str, rows: list[tuple]) -> int:
            """Viết 1 section bảng thống kê, trả về row tiếp theo."""
            ws.cell(row=start_row, column=1).value     = title
            ws.cell(row=start_row, column=1).font      = SECT_FONT
            r = start_row + 1
            # header
            for col, h in enumerate(["Tên", "Số bài", "%"], 1):
                c = ws.cell(row=r, column=col)
                c.value = h; c.fill = HEADER_BG; c.font = HEADER_FONT; c.alignment = CENTER
            r += 1
            total = sum(n for _, n in rows)
            for name, n in rows:
                ws.cell(row=r, column=1).value = name
                ws.cell(row=r, column=2).value = n
                ws.cell(row=r, column=2).fill  = NUM_FILL
                ws.cell(row=r, column=2).alignment = CENTER
                pct = f"{n*100/total:.1f}%" if total else "-"
                ws.cell(row=r, column=3).value = pct
                ws.cell(row=r, column=3).fill  = NUM_FILL
                ws.cell(row=r, column=3).alignment = CENTER
                r += 1
            # tổng cộng
            tc = ws.cell(row=r, column=1); tc.value = "Tổng cộng"; tc.font = Font(bold=True)
            tn = ws.cell(row=r, column=2); tn.value = total; tn.font = Font(bold=True); tn.fill = NUM_FILL; tn.alignment = CENTER
            tp = ws.cell(row=r, column=3); tp.value = "100%"; tp.font = Font(bold=True); tp.fill = NUM_FILL; tp.alignment = CENTER
            return r + 3   # 2 hàng trống sau section

        now_str = datetime.now(_VN_TZ).strftime("%H:%M %d/%m/%Y")
        ws.merge_cells("A1:C1")
        t = ws["A1"]
        t.value = f"📊 TỔNG QUAN — {now_str} | {len(all_data)} bài (24h)"
        t.font  = Font(bold=True, size=14)
        t.alignment = Alignment(horizontal="left")
        ws.row_dimensions[1].height = 24

        # Khu vực
        region_rows = [
            ("🇻🇳 Trong nước", sum(1 for a in all_data if a.get("region")=="vn")),
            ("🌍 Quốc tế",     sum(1 for a in all_data if a.get("region")=="intl")),
        ]
        next_row = _section(3, "📌 Theo khu vực", region_rows)

        # Category
        cat_counter = Counter(a.get("category","") for a in all_data)
        cat_rows = [
            (f"{_CAT_EMOJI.get(k,'')}{CATEGORIES.get(k,k)}", v)
            for k, v in cat_counter.most_common()
        ]
        next_row = _section(next_row, "📂 Theo chủ đề", cat_rows)

        # Source top 20
        src_rows = Counter(a.get("source","") for a in all_data).most_common(20)
        _section(next_row, "📰 Theo nguồn (top 20)", src_rows)

        # Column widths
        for col, w in [(1,28),(2,9),(3,9)]:
            ws.column_dimensions[get_column_letter(col)].width = w

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # xóa sheet mặc định

    today = datetime.now(_VN_TZ).strftime("%d/%m/%Y")

    _write_sheet(wb.create_sheet("🇻🇳 Trong nước"),   vn_arts,   f"🇻🇳 TIN TRONG NƯỚC — {today}  ({len(vn_arts)} bài)")
    _write_sheet(wb.create_sheet("🌍 Quốc tế"),       intl_arts, f"🌍 TIN QUỐC TẾ — {today}  ({len(intl_arts)} bài)")
    _write_sheet(wb.create_sheet("📋 Tất cả"),         arts,      f"📋 TẤT CẢ — {today}  ({len(arts)} bài)")

    ws_sum = wb.create_sheet("📊 Tổng quan")
    _write_summary(ws_sum, arts)
    wb.active = wb["📊 Tổng quan"]   # mở trang tổng quan khi load file

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

async def run_pipeline() -> list[dict]:
    t0 = datetime.now()
    all_articles: list[dict] = []

    # ── Bước 1: RSS ──────────────────────────────────────────────────────────
    rss_sources = INTL_RSS_SOURCES + VN_RSS_SOURCES
    print(f"\n📡  Fetching RSS ({len(rss_sources)} nguồn)...")
    rss_articles = fetch_all_rss(rss_sources)
    print(f"    → {len(rss_articles)} bài RSS")
    all_articles += rss_articles

    # ── Bước 1b: Guardian API ─────────────────────────────────────────────────
    if GUARDIAN_API_KEY:
        print(f"\n📰  Guardian API ({len(_GUARDIAN_SECTIONS)} sections)...")
        g_articles = fetch_guardian_api()
        print(f"    → {len(g_articles)} bài Guardian (full text có sẵn)")
        all_articles += g_articles

    # ── Bước 2: HTML scrape ───────────────────────────────────────────────────
    if not RSS_ONLY:
        print(f"\n🔍  Scraping HTML ({len(VN_SCRAPE_SOURCES)} nguồn)...")
        scrape_articles = fetch_all_scrape(VN_SCRAPE_SOURCES)
        print(f"    → {len(scrape_articles)} bài scrape")
        all_articles += scrape_articles

    # ── Bước 3: Dedup ─────────────────────────────────────────────────────────
    before = len(all_articles)
    unique = deduplicate(all_articles)
    print(f"\n🔄  Dedup: {before} → {len(unique)} bài (bỏ {before - len(unique)} trùng)")

    # ── Bước 4: Enrich ────────────────────────────────────────────────────────
    if NO_ENRICH:
        print("\n⏭️   Skip enrich (--no-enrich)")
        final = [a for a in unique if a.get("title")]
    else:
        print(f"\n⚙️   Enriching {len(unique)} bài ({ENRICH_WORKERS} workers)...")
        final = await enrich_all(unique)

    # ── Bước 5: Lọc 24h ──────────────────────────────────────────────────────
    final, n_dropped = filter_recent(final, hours=FILTER_HOURS)
    print(f"\n🕐  Lọc {FILTER_HOURS}h: giữ {len(final)} bài, bỏ {n_dropped} bài cũ hơn {FILTER_HOURS}h")

    # ── Thống kê ──────────────────────────────────────────────────────────────
    elapsed = (datetime.now() - t0).seconds
    vn_cnt   = sum(1 for a in final if a["region"] == "vn")
    intl_cnt = sum(1 for a in final if a["region"] == "intl")

    print(f"\n{'='*55}")
    print(f"✅  XONG: {len(final)} bài | {elapsed}s ({elapsed//60}p{elapsed%60}s)")
    print(f"    VN={vn_cnt} | QT={intl_cnt}")
    print(f"{'='*55}")

    print("\n📊  Theo category:")
    for cat, count in Counter(a["category"] for a in final).most_common():
        bar   = "█" * (count // 3)
        label = CATEGORIES.get(cat, cat)
        print(f"  {label:30s}: {count:3d}  {bar}")

    if not NO_ENRICH:
        print("\n🔧  Enrich method:")
        for method, count in Counter(a.get("enrich_method", "") for a in final).most_common():
            print(f"  {method:30s}: {count}")

    # Export JSON mẫu — lấy 4 bài từ mỗi enrich_method để đa dạng nguồn
    from collections import defaultdict
    import random
    buckets: dict[str, list] = defaultdict(list)
    for a in final:
        buckets[a.get("enrich_method", "")].append(a)
    sample: list = []
    for method_arts in buckets.values():
        sample.extend(random.sample(method_arts, min(4, len(method_arts))))
    out_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "articles_v2_output.json")
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(sample[:30], f, ensure_ascii=False, indent=2)
    print(f"\n📄  Export {len(sample[:30])} bài mẫu (đa dạng method) → articles_v2_output.json")

    # ── Excel export — master file tích lũy ─────────────────────────────────
    master_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), MASTER_EXCEL_FILE)
    n_new = append_to_master_excel(final, master_path)
    print(f"📊  Excel: +{n_new} bài mới → {MASTER_EXCEL_FILE}  (tổng {len(final)} bài lần chạy này)")

    return final


async def fetch_articles(
    hours: int = 24,
    rss_only: bool = False,
    no_enrich: bool = False,
) -> list[dict]:
    """Public API — chạy pipeline, trả về list[dict], không export file.
    Dùng bởi morning_run.py thay thế scraper.py.
    """
    all_articles: list[dict] = []

    rss_sources = INTL_RSS_SOURCES + VN_RSS_SOURCES
    all_articles += fetch_all_rss(rss_sources)

    if GUARDIAN_API_KEY:
        all_articles += fetch_guardian_api()

    if not rss_only:
        all_articles += fetch_all_scrape(VN_SCRAPE_SOURCES)

    unique = deduplicate(all_articles)

    if no_enrich:
        final = [a for a in unique if a.get("title")]
    else:
        final = await enrich_all(unique)

    final, _ = filter_recent(final, hours=hours)
    return final


if __name__ == "__main__":
    asyncio.run(run_pipeline())
